"""
Enhanced Alpha Vantage News Generator
- Professional, dynamic content style with varied transitions
- Motion script generation for video production
- Episode title generation replacing ticker columns
- Voice-friendly content for HeyGen compatibility
"""

import requests
import json
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
import os
from dotenv import load_dotenv
from openai import OpenAI
import time

# Load environment variables
load_dotenv()


class ProfessionalNewsGenerator:
    def __init__(self, debug_mode=False):
        self.api_key = os.getenv('ALPHA_VANTAGE_API_KEY')
        self.openai_key = os.getenv('OPENAI_API_KEY')
        self.debug_mode = debug_mode

        if not self.api_key:
            raise ValueError("ALPHA_VANTAGE_API_KEY required in .env file")
        if not self.openai_key:
            raise ValueError("OPENAI_API_KEY required in .env file")

        self.base_url = "https://www.alphavantage.co/query"
        self.openai_client = OpenAI(api_key=self.openai_key)
        self.call_count = 0

    # ===== Main Flow =====
    def generate_content(self) -> Dict[str, Any]:
        today = datetime.now().strftime('%A')
        print(f"📅 Generating professional content for {today}")

        if today == 'Monday':
            return self._generate_monday()
        elif today == 'Wednesday':
            return self._generate_wednesday()
        elif today == 'Friday':
            return self._generate_friday()
        else:
            return self._generate_generic()

    # ===== Day Modes =====
    def _generate_monday(self):
        news = self._get_high_quality_news(limit=8)
        market_data = self._get_market_snapshot()
        script, social_post, motion_script, video_caption, episode_title = self._generate_content(
            news, "Monday", "Weekly market open with key developments", 60
        )
        return {
            'day': 'Monday',
            'type': 'Weekly Market Open',
            'script': script,
            'social_post': social_post,
            'motion_script': motion_script,
            'video_caption': video_caption,
            'episode_title': episode_title,
            'news_count': len(news),
            'market_data': market_data,
            'top_articles': news[:3]
        }

    def _generate_wednesday(self):
        news = self._get_high_quality_news_timeframe(days_back=2, limit=10)
        movers = self._get_recent_movers()
        script, social_post, motion_script, video_caption, episode_title = self._generate_content(
            news, "Wednesday", "Mid-week market analysis", 75
        )
        return {
            'day': 'Wednesday',
            'type': 'Mid-Week Analysis',
            'script': script,
            'social_post': social_post,
            'motion_script': motion_script,
            'video_caption': video_caption,
            'episode_title': episode_title,
            'news_count': len(news),
            'top_movers': movers,
            'top_articles': news[:3]
        }

    def _generate_friday(self):
        news = self._get_high_quality_news_timeframe(days_back=5, limit=12)
        weekly_summary = self._get_weekly_summary()
        script, social_post, motion_script, video_caption, episode_title = self._generate_content(
            news, "Friday", "Weekly market wrap-up", 90
        )
        return {
            'day': 'Friday',
            'type': 'Weekly Market Wrap',
            'script': script,
            'social_post': social_post,
            'motion_script': motion_script,
            'video_caption': video_caption,
            'episode_title': episode_title,
            'news_count': len(news),
            'weekly_summary': weekly_summary,
            'top_articles': news[:3]
        }

    def _generate_generic(self):
        news = self._get_high_quality_news(limit=8)
        market_data = self._get_market_snapshot()
        script, social_post, motion_script, video_caption, episode_title = self._generate_content(
            news, datetime.now().strftime('%A'), "Daily market update", 60
        )
        return {
            'day': datetime.now().strftime('%A'),
            'type': 'Daily Market Update',
            'script': script,
            'social_post': social_post,
            'motion_script': motion_script,
            'video_caption': video_caption,
            'episode_title': episode_title,
            'news_count': len(news),
            'market_data': market_data,
            'top_articles': news[:3]
        }

    # ===== Content Generation =====
    def _generate_content(self, news: List[Dict], day: str, theme: str, target_seconds: int) -> Tuple[str, str, str, str, str]:
        """Generate professional, dynamic content with varied transitions"""
        
        if not news:
            fallback_script = f"Here's your market update for {datetime.now().strftime('%B %d')}. First up — markets are showing consolidation today with investors watching for key developments. Next — the Federal Reserve continues monitoring economic indicators closely. And finally — several major companies are preparing quarterly results this week. That's your {datetime.now().strftime('%B %d')} rundown — see you tomorrow!"
            fallback_social = f"📊 Markets in consolidation mode today. Fed watching indicators closely. Major earnings ahead. #Markets #Trading #Finance"
            fallback_motion = "Start with confident eye contact. Light gesture on 'consolidation'. Authoritative posture on Fed mention. End with professional nod."
            fallback_caption = f"Daily Market Update - {datetime.now().strftime('%B %d, %Y')} | Consolidation & Fed Watch"
            fallback_title = f"Market Update - {datetime.now().strftime('%B %d')} | Markets, Fed, and Earnings"
            return fallback_script, fallback_social, fallback_motion, fallback_caption, fallback_title

        # Prepare context with top stories
        top_stories = []
        for i, article in enumerate(news[:5], 1):  # Use top 5 stories for more context
            tickers_str = f" (${', '.join(article['tickers'][:2])})" if article.get('tickers') else ""
            top_stories.append(f"{i}. {article['title']}{tickers_str}")
            top_stories.append(f"   Sentiment: {article['sentiment']} ({article['sentiment_score']:.2f})")
            if article.get('summary'):
                top_stories.append(f"   Summary: {article['summary'][:100]}...")
        
        context = f"Date: {datetime.now().strftime('%B %d, %Y')}\nTop Stories Available:\n" + "\n".join(top_stories)
        
        # Professional system prompt with voice-friendly guidelines
        system_prompt = """You are a professional financial news presenter creating concise, authoritative market updates. Your style should match financial news broadcasts - direct, informative, and professional.

STYLE GUIDELINES:
- Start with "Here's your market update for [date]"
- Present 2-3 key stories with varied, natural transitions
- Use professional language: "analysts warn", "sector continues to", "drawing bullish sentiment"
- Include specific numbers, projections, and company names when available
- Keep sentences clear and declarative
- End with a brief, professional sign-off
- Sound authoritative but accessible

TICKER SYMBOL RULES FOR VOICE SYNTHESIS:
- NEVER write ticker symbols that will be read aloud (e.g., $AAPL, FOREX:USD, CRYPTO:BTC)
- Instead use company/asset names: "Apple", "US Dollar", "Bitcoin"
- For unknown tickers, use generic terms: "the company", "related stocks", "cryptocurrency markets"
- Only use ticker symbols in parentheses if they add clarity: "Apple (AAPL)"
- Avoid complex symbols like FOREX:KRW, CRYPTO:BTC - say "Korean Won", "Bitcoin" instead

CONTENT STRUCTURE:
1. Date-specific opening
2. Three stories with varied transitions
3. Market context and implications
4. Professional closing

CRITICAL: Return ONLY valid JSON with script, social, motion, caption, and title keys."""

        user_prompt = f"""Create a professional {target_seconds}-second financial news script in the style of the sample below:

SAMPLE STYLE:
"Here's your market update for August 20. First up — a Fed official says staff should be allowed to hold some crypto. That's a potential boost for Ethereum and related markets.
Next — High Arctic is making executive management changes. Investors are watching closely since leadership shifts often signal strategy updates.
And finally — Bitcoin's under pressure. The big question on the street: is $112K the final bottom?
That's your August 20 rundown — see you tomorrow!"

CONTEXT:
{context}

Return your answer in JSON format exactly like this:
{{
  "script": "professional news script here",
  "social": "280-character social post here", 
  "motion": "300-character motion direction here",
  "caption": "video caption here",
  "title": "episode title here"
}}

SCRIPT REQUIREMENTS:
- {target_seconds} seconds when spoken at news pace (~{target_seconds * 2.5} words)
- EXACTLY 3 STORIES with varied, natural transitions
- Start with "Here's your market update for [current date]"
- Use VARIED transition phrases (mix and match):
  * Story 1: "First up", "Leading off", "Starting with", "Top story"
  * Story 2: "Next", "Meanwhile", "Moving on", "Also", "In other news"  
  * Story 3: "And finally", "Lastly", "Also worth noting", "And", "Wrapping up"
- Add market context and implications for each story
- Include phrases like "That's a potential boost for...", "Investors are watching...", "The big question is..."
- End with "That's your [date] rundown — see you tomorrow!"
- Sound conversational but professional
- VOICE-FRIENDLY: Use company names, not ticker symbols (say "Apple" not "$AAPL")
- For crypto: say "Bitcoin", "Ethereum", not "CRYPTO:BTC"
- For forex: say "US Dollar", "Korean Won", not "FOREX:USD"
- For unknown tickers: use "the company", "related stocks", "the sector"
- EXCLUDE: Legal notices, shareholder alerts, class action lawsuits, attorney notices
- FOCUS: Market movements, earnings, economic policy, sector trends, corporate developments
- TARGET LENGTH: Aim for {target_seconds * 2.5} words minimum for proper pacing

TITLE REQUIREMENTS:
- Format: "Market Update - [Date] | [3 key themes/sectors]"
- Example: "Market Update - August 20 | Crypto, Energy, and Bitcoin"
- Use 3 concise sector/theme words separated by commas and "and"
- Keep it under 60 characters total
- Focus on the main sectors/themes covered in the script

SOCIAL POST REQUIREMENTS:
- Professional summary of key developments
- Include 1-2 relevant tickers or numbers
- 280 characters max, 1 emoji, 2-3 hashtags
- News-focused, not promotional

MOTION SCRIPT REQUIREMENTS (300 characters max):
- Professional presenter body language directions
- Include: opening stance, key gesture moments, transitions, closing
- Specify eye contact, hand gestures, posture changes
- Match the tone and key points of the script

VIDEO CAPTION REQUIREMENTS:
- Professional title for the video
- Include date and key topics covered
- Suitable for video thumbnail/title
- 60-80 characters ideal

THEME: {theme}

IMPORTANT: Only use legitimate market news, economic developments, earnings reports, and corporate announcements. Ignore legal notices, shareholder alerts, and lawsuit announcements."""
        
        try:
            print(f"🤖 Generating professional content with GPT-4o...")
            
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                max_completion_tokens=1200,
                temperature=0.3,  # Lower temperature for more consistent professional tone
                response_format={"type": "json_object"}
            )
            
            full_output = response.choices[0].message.content.strip()
            
            # Parse JSON response
            try:
                parsed = json.loads(full_output)
            except json.JSONDecodeError as e:
                print(f"❌ JSON parsing error: {e}")
                print(f"Raw output: {full_output[:200]}...")
                return self._create_fallback_content(news, day)

            script = parsed.get("script", "").strip()
            social = parsed.get("social", "").strip()
            motion = parsed.get("motion", "").strip()
            caption = parsed.get("caption", "").strip()
            title = parsed.get("title", "").strip()

            # Quality validation
            if self._validate_professional_content(script, social, motion, caption, target_seconds):
                print(f"✅ Success with GPT-4o")
                return script, social, motion, caption, title
            else:
                print(f"⚠️ Content quality check failed, using fallback")
                return self._create_fallback_content(news, day)

        except Exception as e:
            print(f"❌ Error generating content: {str(e)}")
            return self._create_fallback_content(news, day)

    def _validate_professional_content(self, script: str, social: str, motion: str, caption: str, target_seconds: int) -> bool:
        """Validate generated content meets professional standards"""
        
        # Check script structure
        if not script.startswith("Here's your market update"):
            return False
        
        # Check for minimum 3 stories with varied transitions
        transition_words = ["first up", "next", "and finally", "meanwhile", "leading off", "moving on", "lastly", "also", "wrapping up"]
        found_transitions = sum(1 for word in transition_words if word in script.lower())
        if found_transitions < 2:  # Should have at least 2 transitions for 3 stories
            return False
        
        # Check minimum length (professional pace is ~2.5 words/second)
        min_words = max(150, target_seconds * 2.2)  # Increased minimum
        if len(script.split()) < min_words:
            return False
        
        # Check social post length
        if len(social) > 280 or len(social) < 50:
            return False
        
        # Check motion script length
        if len(motion) > 300 or len(motion) < 50:
            return False
        
        # Check caption length
        if len(caption) > 100 or len(caption) < 30:
            return False
        
        return True

    def _create_fallback_content(self, news: List[Dict], day: str) -> Tuple[str, str, str, str, str]:
        """Professional fallback content with voice-friendly formatting and dynamic transitions"""
        current_date = datetime.now().strftime('%B %d')
        
        if not news:
            script = f"Here's your market update for {current_date}. First up — markets are showing mixed signals today as investors digest recent economic data and await key earnings reports. Next — the Federal Reserve continues to monitor inflation indicators closely, with analysts expecting potential policy adjustments in the coming weeks. And finally — several major companies are reporting quarterly results this week, setting the tone for sector performance. That's your {current_date} rundown — see you tomorrow!"
            social = f"📊 Mixed market signals on {current_date} as investors digest economic data. Fed monitoring inflation closely. Major earnings this week. #Markets #Trading #Finance"
            motion = "Professional opening with steady eye contact. Light gesture on 'mixed signals'. Authoritative posture on Fed mention. End with confident nod."
            caption = f"Market Update - {current_date} | Mixed Signals, Fed Watch & Earnings"
            title = f"Market Update - {current_date} | Markets, Fed, and Earnings"
            return script, social, motion, caption, title
        
        # Build comprehensive script with varied transitions
        script_parts = [f"Here's your market update for {current_date}."]
        
        # Story 1 - varied opening transitions
        openings = ["First up", "Leading off", "Starting with", "Top story"]
        script_parts.append(f"{openings[0]} — {news[0]['title'].lower()}.")
        
        # Add context for first story
        if news[0].get('tickers'):
            voice_friendly_tickers = self._convert_tickers_for_voice(news[0]['tickers'][:2])
            if voice_friendly_tickers:
                script_parts.append(f"That's a potential boost for {', '.join(voice_friendly_tickers)} and related markets.")
        
        # Story 2
        if len(news) > 1:
            transitions = ["Next", "Meanwhile", "Moving on", "Also", "In other news"]
            script_parts.append(f"{transitions[0]} — {news[1]['title'].lower()}.")
            script_parts.append("Investors are watching closely since leadership shifts often signal strategy updates.")
        
        # Story 3
        if len(news) > 2:
            endings = ["And finally", "Lastly", "Also worth noting", "Wrapping up"]
            script_parts.append(f"{endings[0]} — {news[2]['title'].lower()}.")
            script_parts.append("The big question on the street: what's the next move?")
        else:
            # Generic third story if no more news available
            script_parts.append("And finally — market volatility continues as traders position for upcoming economic data.")
            script_parts.append("The big question: where do we go from here?")
        
        script_parts.append(f"That's your {current_date} rundown — see you tomorrow!")
        
        script = " ".join(script_parts)
        
        # Generate title based on story themes
        themes = []
        if news:
            for article in news[:3]:
                title_lower = article['title'].lower()
                if any(word in title_lower for word in ['crypto', 'bitcoin', 'ethereum']):
                    themes.append("Crypto")
                elif any(word in title_lower for word in ['fed', 'federal', 'interest']):
                    themes.append("Fed")
                elif any(word in title_lower for word in ['energy', 'oil', 'gas']):
                    themes.append("Energy")
                elif any(word in title_lower for word in ['earnings', 'profit', 'revenue']):
                    themes.append("Earnings")
                elif any(word in title_lower for word in ['tech', 'technology', 'ai']):
                    themes.append("Tech")
                else:
                    themes.append("Markets")
        
        # Remove duplicates and limit to 3
        unique_themes = list(dict.fromkeys(themes))[:3]
        if len(unique_themes) < 3:
            unique_themes.extend(["Markets"] * (3 - len(unique_themes)))
        
        title = f"Market Update - {current_date} | {', '.join(unique_themes[:-1])}, and {unique_themes[-1]}"
        
        # Keep tickers in social post for visual reference
        tickers_text = f"${'/'.join(news[0]['tickers'][:2])}" if news and news[0].get('tickers') else "key sectors"
        social = f"📈 {news[0]['title'][:100] if news else 'Market developments'}... {tickers_text} in focus. Multiple stories across markets today. #Markets #Trading #Finance"
        
        motion = f"Start with confident eye contact. Emphasize transitions with slight gestures. Maintain professional posture. End with authoritative nod and slight smile."
        
        caption = f"Market Update - {current_date} | {unique_themes[0]}, {unique_themes[1]} & More"
        
        return script, social, motion[:300], caption, title

    def _convert_tickers_for_voice(self, tickers: List[str]) -> List[str]:
        """Convert ticker symbols to voice-friendly company/asset names"""
        ticker_map = {
            # Major stocks
            'AAPL': 'Apple',
            'MSFT': 'Microsoft', 
            'GOOGL': 'Google',
            'AMZN': 'Amazon',
            'TSLA': 'Tesla',
            'NVDA': 'NVIDIA',
            'META': 'Meta',
            'NFLX': 'Netflix',
            'GRMN': 'Garmin',
            
            # Crypto patterns
            'BTC': 'Bitcoin',
            'ETH': 'Ethereum',
            'CRYPTO:BTC': 'Bitcoin',
            'CRYPTO:ETH': 'Ethereum',
            
            # Forex patterns  
            'FOREX:USD': 'US Dollar',
            'FOREX:EUR': 'Euro',
            'FOREX:GBP': 'British Pound',
            'FOREX:JPY': 'Japanese Yen',
            'FOREX:KRW': 'Korean Won',
            'FOREX:CNY': 'Chinese Yuan',
            
            # Common unknown patterns
            'TOP': 'the company',
            'UNKNOWN': 'related stocks'
        }
        
        voice_friendly = []
        for ticker in tickers:
            if ticker in ticker_map:
                voice_friendly.append(ticker_map[ticker])
            elif ticker.startswith('FOREX:'):
                # Handle any FOREX: pattern generically
                currency = ticker.replace('FOREX:', '')
                voice_friendly.append(f"the {currency}")
            elif ticker.startswith('CRYPTO:'):
                # Handle any CRYPTO: pattern generically  
                crypto = ticker.replace('CRYPTO:', '')
                voice_friendly.append(f"{crypto} cryptocurrency")
            elif len(ticker) <= 4 and ticker.isupper():
                # Likely a stock ticker
                voice_friendly.append("the company")
            else:
                # Unknown format
                voice_friendly.append("related markets")
        
        return voice_friendly

    def _calculate_market_relevance_score(self, article) -> float:
        """Calculate a relevance score based on market impact factors"""
        score = 0.0
        title = article['title'].lower()
        summary = article.get('summary', '').lower()
        source = article.get('source', '').lower()
        
        # 1. Source Credibility (0-25 points)
        premium_sources = [
            'reuters', 'bloomberg', 'cnbc', 'marketwatch', 'wsj', 'financial times',
            'barrons', 'seeking alpha', 'yahoo finance', 'benzinga', 'zacks',
            'motley fool', 'investing.com', 'nasdaq', 'sec.gov'
        ]
        if any(src in source for src in premium_sources):
            score += 25
        elif any(src in source for src in ['prnewswire', 'businesswire', 'globe newswire']):
            score += 10  # Press releases get some points but less
        
        # 2. Market Impact Keywords (0-30 points)
        high_impact_keywords = [
            'fed', 'federal reserve', 'interest rate', 'inflation', 'gdp', 'unemployment',
            'earnings', 'profit', 'revenue', 'guidance', 'outlook', 'forecast',
            'merger', 'acquisition', 'ipo', 'stock split', 'dividend',
            'sec', 'regulation', 'antitrust', 'approval', 'fda',
            'oil price', 'gold', 'dollar', 'treasury', 'bond yield'
        ]
        
        medium_impact_keywords = [
            'ceo', 'cfo', 'executive', 'leadership', 'appointment',
            'partnership', 'deal', 'contract', 'launch', 'product',
            'upgrade', 'downgrade', 'analyst', 'price target',
            'bitcoin', 'crypto', 'blockchain', 'ai', 'artificial intelligence'
        ]
        
        high_matches = sum(1 for keyword in high_impact_keywords if keyword in title or keyword in summary)
        medium_matches = sum(1 for keyword in medium_impact_keywords if keyword in title or keyword in summary)
        
        score += min(high_matches * 10, 30)  # Max 30 points for high impact
        score += min(medium_matches * 5, 15)   # Max 15 points for medium impact
        
        # 3. Ticker Presence and Quality (0-20 points)
        tickers = article.get('tickers', [])
        if tickers:
            # Major market cap companies get higher scores
            major_tickers = [
                'AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'NVDA', 'META', 'NFLX',
                'SPY', 'QQQ', 'BTC', 'ETH', 'JPM', 'BAC', 'WMT', 'JNJ', 'PG'
            ]
            major_count = sum(1 for ticker in tickers if ticker in major_tickers)
            score += min(major_count * 8, 20)  # 8 points per major ticker, max 20
            
            if not major_count and len(tickers) > 0:
                score += 5  # Some points for having any tickers
        
        # 4. Sentiment Strength (0-15 points)
        sentiment_score = abs(article.get('sentiment_score', 0))
        if sentiment_score > 0.5:
            score += 15  # Strong sentiment (positive or negative)
        elif sentiment_score > 0.3:
            score += 10  # Moderate sentiment
        elif sentiment_score > 0.1:
            score += 5   # Mild sentiment
        
        # 5. Recency Factor (0-10 points)
        try:
            if article.get('time_published'):
                # Parse timestamp and give points for recency
                from datetime import datetime
                pub_time = datetime.strptime(article['time_published'], '%Y%m%dT%H%M%S')
                now = datetime.now()
                hours_old = (now - pub_time).total_seconds() / 3600
                
                if hours_old < 2:
                    score += 10  # Very recent
                elif hours_old < 6:
                    score += 8   # Recent
                elif hours_old < 12:
                    score += 5   # Same day
                elif hours_old < 24:
                    score += 2   # Yesterday
        except:
            score += 5  # Default for unparseable dates
        
        # 6. Title Quality (0-10 points)
        title_length = len(title.split())
        if 8 <= title_length <= 20:  # Optimal title length
            score += 10
        elif 5 <= title_length <= 25:
            score += 5
        
        # Avoid clickbait patterns
        clickbait_patterns = ['you won\'t believe', 'shocking', 'this will', 'must see']
        if any(pattern in title for pattern in clickbait_patterns):
            score -= 20
        
        return round(score, 1)

    # ===== News Fetch with Quality Scoring =====
    def _get_high_quality_news(self, limit=8):
        """Get high-quality, market-relevant news with intelligent scoring"""
        params = {
            'function': 'NEWS_SENTIMENT',
            'apikey': self.api_key,
            'limit': 100,  # Get more articles to filter from
            'sort': 'RELEVANCE'  # Changed from LATEST to RELEVANCE
        }
        try:
            print(f"🔍 Calling Alpha Vantage API for market-relevant news...")
            response = requests.get(self.base_url, params=params, timeout=30)
            self.call_count += 1
            
            print(f"📡 API Response Status: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                print(f"📊 API Response Keys: {list(data.keys())}")
                
                if 'feed' in data:
                    raw_articles = data['feed']
                    print(f"📰 Raw articles found: {len(raw_articles)}")
                    
                    # Score and filter articles
                    scored_articles = []
                    for article in raw_articles:
                        if self._is_basic_financial_content(article):
                            normalized = self._normalize_article(article)
                            quality_score = self._calculate_market_relevance_score(normalized)
                            normalized['quality_score'] = quality_score
                            scored_articles.append(normalized)
                    
                    # Sort by quality score (highest first)
                    scored_articles.sort(key=lambda x: x['quality_score'], reverse=True)
                    
                    if self.debug_mode:
                        print(f"📋 High-quality articles after scoring: {len(scored_articles)}")
                        if scored_articles:
                            print("\n🏆 Top scored articles:")
                            for i, article in enumerate(scored_articles[:5]):
                                print(f"{i+1}. Score: {article['quality_score']:.1f} - {article['title'][:80]}...")
                                print(f"   Source: {article['source']} | Sentiment: {article['sentiment']} ({article['sentiment_score']:.2f})")
                                print(f"   Tickers: {', '.join(article['tickers'][:3]) if article['tickers'] else 'None'}")
                    
                    return scored_articles[:limit]
                else:
                    print(f"❌ No 'feed' key in response. Available keys: {list(data.keys())}")
                    if 'Error Message' in data:
                        print(f"❌ API Error: {data['Error Message']}")
                    if 'Note' in data:
                        print(f"ℹ️ API Note: {data['Note']}")
            else:
                print(f"❌ API error: {response.status_code}")
                print(f"Response text: {response.text[:200]}...")
                
        except Exception as e:
            print(f"❌ Error getting news: {e}")
        return []

    def _get_high_quality_news_timeframe(self, days_back, limit):
        """Get quality news within a specific timeframe with scoring"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days_back)
        params = {
            'function': 'NEWS_SENTIMENT',
            'apikey': self.api_key,
            'limit': 100,  # Get more to filter from
            'time_from': start_date.strftime('%Y%m%dT0000'),
            'time_to': end_date.strftime('%Y%m%dT2359'),
            'sort': 'RELEVANCE'  # Prioritize relevance over recency
        }
        try:
            response = requests.get(self.base_url, params=params, timeout=30)
            self.call_count += 1
            if response.status_code == 200:
                data = response.json()
                if 'feed' in data:
                    raw_articles = data['feed']
                    
                    # Score and filter articles
                    scored_articles = []
                    for article in raw_articles:
                        if self._is_basic_financial_content(article):  # Use the correct method name
                            normalized = self._normalize_article(article)
                            quality_score = self._calculate_market_relevance_score(normalized)
                            normalized['quality_score'] = quality_score
                            scored_articles.append(normalized)
                    
                    # Sort by quality score
                    scored_articles.sort(key=lambda x: x['quality_score'], reverse=True)
                    return scored_articles[:limit]
        except Exception as e:
            print(f"❌ Error getting timeframe news: {e}")
        return []

    def _is_basic_financial_content(self, article):
        """Enhanced filtering for high-quality financial content only"""
        title = article.get('title', '').lower()
        summary = article.get('summary', '').lower()
        source = article.get('source', '').lower()
        
        # Immediate exclusions - these are never quality financial news
        exclude_terms = [
            'horoscope', 'recipe', 'dating', 'celebrity gossip', 'weather forecast',
            'shareholder alert', 'class action lawsuit', 'lead plaintiff deadline',
            'former attorney general', 'law firm', 'legal notice', 'litigation',
            'lawsuit against', 'reminds investors', 'kahn swick', 'losses in excess',
            'securities fraud', 'ponzi scheme', 'pump and dump',
            'weight loss', 'diet pill', 'miracle cure', 'get rich quick',
            'earn money from home', 'work from home scam'
        ]
        
        if any(term in title or term in summary for term in exclude_terms):
            return False
        
        # Exclude ALL CAPS titles (usually spam)
        if title.isupper() and len(title) > 20:
            return False
        
        # Exclude obvious promotional content
        promotional_phrases = [
            'paid promotion', 'sponsored content', 'advertisement',
            'buy now', 'limited time offer', 'act fast'
        ]
        if any(phrase in title or phrase in summary for phrase in promotional_phrases):
            return False
        
        # Must have financial relevance
        financial_indicators = [
            # Core finance terms
            'stock', 'market', 'trading', 'investment', 'earnings', 'revenue',
            'profit', 'share', 'price', 'analyst', 'economy', 'economic',
            'financial', 'nasdaq', 'dow', 's&p', 'fed', 'inflation', 'gdp',
            
            # Crypto and digital assets
            'crypto', 'bitcoin', 'ethereum', 'blockchain', 'defi',
            
            # Corporate terms
            'merger', 'acquisition', 'ipo', 'ceo', 'cfo', 'board',
            'dividend', 'split', 'buyback', 'guidance',
            
            # Market sectors
            'bank', 'rates', 'bonds', 'commodities', 'oil', 'gold',
            'tech', 'healthcare', 'energy', 'utilities', 'reit'
        ]
        
        has_financial_terms = any(term in title or term in summary for term in financial_indicators)
        has_tickers = len(article.get('ticker_sentiment', [])) > 0
        
        # Quality length check
        is_reasonable_length = 20 <= len(title) <= 200
        
        # Must meet minimum criteria
        return (has_financial_terms or has_tickers) and is_reasonable_length

    def _normalize_article(self, article):
        tickers = [item.get('ticker') for item in article.get('ticker_sentiment', []) if item.get('ticker')]
        
        return {
            'title': article.get('title', ''),
            'summary': article.get('summary', ''),
            'source': article.get('source', ''),
            'time_published': article.get('time_published', ''),
            'sentiment': article.get('overall_sentiment_label', 'neutral'),
            'sentiment_score': float(article.get('overall_sentiment_score', 0)),
            'tickers': tickers,
            'url': article.get('url', ''),
            'quality_score': 75
        }

    # ===== Market Data =====
    def _get_market_snapshot(self):
        try:
            import yfinance as yf
            spy = yf.Ticker('SPY')
            hist = spy.history(period="2d")
            if len(hist) >= 2:
                current, previous = hist['Close'].iloc[-1], hist['Close'].iloc[-2]
                change_pct = ((current - previous) / previous) * 100
                return {
                    'symbol': 'SPY', 
                    'price': f"${current:.2f}", 
                    'change': f"{change_pct:+.2f}%", 
                    'status': 'connected'
                }
        except ImportError:
            print("⚠️ yfinance not installed. Install with: pip install yfinance")
        except Exception as e:
            print(f"❌ Error getting market snapshot: {e}")
        return {'status': 'error'}

    def _get_recent_movers(self):
        try:
            import yfinance as yf
            symbols = ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA', 'NVDA']
            movers = []
            for symbol in symbols[:4]:
                ticker = yf.Ticker(symbol)
                hist = ticker.history(period="2d")
                if len(hist) >= 2:
                    current, previous = hist['Close'].iloc[-1], hist['Close'].iloc[-2]
                    change_pct = ((current - previous) / previous) * 100
                    if abs(change_pct) > 1.5:
                        movers.append({
                            'symbol': symbol, 
                            'change': f"{change_pct:+.2f}%", 
                            'price': f"${current:.2f}"
                        })
            return movers
        except Exception as e:
            print(f"❌ Error getting movers: {e}")
            return []

    def _get_weekly_summary(self):
        try:
            import yfinance as yf
            spy = yf.Ticker('SPY')
            hist = spy.history(period="1mo")
            if len(hist) >= 5:
                current, week_ago = hist['Close'].iloc[-1], hist['Close'].iloc[-5]
                weekly_change = ((current - week_ago) / week_ago) * 100
                week_high = hist['High'].tail(5).max()
                week_low = hist['Low'].tail(5).min()
                return {
                    'weekly_change': f"{weekly_change:+.1f}%", 
                    'week_high': f"${week_high:.2f}", 
                    'week_low': f"${week_low:.2f}"
                }
        except Exception as e:
            print(f"❌ Error getting weekly summary: {e}")
        return {}

    # ===== Excel Saving with Episode Title =====
    def save_to_excel(self, content: Dict[str, Any]) -> str:
        """Save content to Excel with episode title replacing top tickers"""
        filename = "market_content_tracker.xlsx"
        
        try:
            file_exists = os.path.exists(filename)
            
            if file_exists:
                try:
                    existing_df = pd.read_excel(filename, sheet_name='Content_Log')
                    print(f"📄 Updating existing file: {filename}")
                    print(f"📊 Current entries in file: {len(existing_df)}")
                except Exception as e:
                    print(f"⚠️ Could not read existing file ({e}), creating new structure")
                    existing_df = pd.DataFrame(columns=[
                        'Date', 'Time', 'Day', 'Content_Type', 'Script', 'Social_Post',
                        'Motion_Script', 'Video_Caption', 'Episode_Title', 'Script_Length', 'Word_Count', 
                        'News_Count', 'Market_Data', 'Quality_Score'
                    ])
            else:
                existing_df = pd.DataFrame(columns=[
                    'Date', 'Time', 'Day', 'Content_Type', 'Script', 'Social_Post',
                    'Motion_Script', 'Video_Caption', 'Episode_Title', 'Script_Length', 'Word_Count', 
                    'News_Count', 'Market_Data', 'Quality_Score'
                ])
                print(f"📄 Creating new file: {filename}")
            
            # Prepare new row data
            current_time = datetime.now()
            
            market_data_str = ""
            if content.get('market_data') and content['market_data'].get('status') == 'connected':
                market_data_str = f"SPY: {content['market_data']['price']} ({content['market_data']['change']})"
            
            new_row = {
                'Date': current_time.strftime('%Y-%m-%d'),
                'Time': current_time.strftime('%H:%M:%S'),
                'Day': content['day'],
                'Content_Type': content['type'],
                'Script': content['script'],
                'Social_Post': content['social_post'],
                'Motion_Script': content.get('motion_script', ''),
                'Video_Caption': content.get('video_caption', ''),
                'Episode_Title': content.get('episode_title', ''),
                'Script_Length': len(content['script']),
                'Word_Count': len(content['script'].split()),
                'News_Count': content.get('news_count', 0),
                'Market_Data': market_data_str,
                'Quality_Score': 'Generated' if content.get('news_count', 0) > 0 else 'Fallback'
            }
            
            new_row_df = pd.DataFrame([new_row])
            new_df = pd.concat([existing_df, new_row_df], ignore_index=True)
            
            print(f"📝 Adding new row: {new_row['Date']} {new_row['Time']} - {new_row['Day']}")
            print(f"📄 Script preview: {new_row['Script'][:50]}...")
            print(f"🎬 Motion script: {new_row['Motion_Script'][:50]}...")
            print(f"📺 Video caption: {new_row['Video_Caption']}")
            print(f"🎯 Episode title: {new_row['Episode_Title']}")
            
            # Sort by Date and Time (most recent first)
            new_df['DateTime'] = pd.to_datetime(new_df['Date'] + ' ' + new_df['Time'])
            new_df = new_df.sort_values('DateTime', ascending=False).drop('DateTime', axis=1)
            
            print(f"📊 Total rows to save: {len(new_df)}")
            
            # Save to Excel
            print("💾 Saving to Excel...")
            with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
                new_df.to_excel(writer, sheet_name='Content_Log', index=False, header=True)
                print(f"📋 Content_Log sheet saved with {len(new_df)} rows")
                
                # Create summary sheet
                summary_data = pd.DataFrame({
                    'Metric': [
                        'Total Entries', 'Last Updated', 'Scripts This Week', 
                        'Most Active Day', 'Average Script Length', 'Total Words Generated'
                    ],
                    'Value': [
                        len(new_df),
                        current_time.strftime('%Y-%m-%d %H:%M:%S'),
                        len(new_df[new_df['Date'] >= (current_time - timedelta(days=7)).strftime('%Y-%m-%d')]),
                        new_df['Day'].mode().iloc[0] if len(new_df) > 0 else 'None',
                        f"{new_df['Script_Length'].mean():.0f} characters" if len(new_df) > 0 else '0',
                        new_df['Word_Count'].sum() if len(new_df) > 0 else 0
                    ]
                })
                summary_data.to_excel(writer, sheet_name='Summary', index=False)
                print("📈 Summary sheet saved")
                
                # Today's articles (if available)
                if content.get('top_articles'):
                    articles_data = []
                    for i, article in enumerate(content['top_articles'], 1):
                        articles_data.append({
                            'Rank': i,
                            'Title': article['title'],
                            'Source': article['source'],
                            'Sentiment': article['sentiment'],
                            'Sentiment_Score': article['sentiment_score'],
                            'Tickers': ', '.join(article.get('tickers', [])),
                            'Time_Published': article['time_published']
                        })
                    
                    pd.DataFrame(articles_data).to_excel(writer, sheet_name='Todays_Sources', index=False)
                    print("📰 Today's Sources sheet saved")
            
            print("💾 Excel file write completed")
            
            # Verify the file was written correctly
            try:
                verify_df = pd.read_excel(filename, sheet_name='Content_Log')
                print(f"✅ Verification: File contains {len(verify_df)} rows")
                if len(verify_df) > 0:
                    print(f"📝 Latest entry: {verify_df.iloc[0]['Date']} - {verify_df.iloc[0]['Day']}")
                    print(f"🎬 Motion script included: {'Yes' if verify_df.iloc[0].get('Motion_Script') else 'No'}")
                    print(f"📺 Video caption included: {'Yes' if verify_df.iloc[0].get('Video_Caption') else 'No'}")
                    print(f"🎯 Episode title included: {'Yes' if verify_df.iloc[0].get('Episode_Title') else 'No'}")
            except Exception as e:
                print(f"⚠️ Verification failed: {e}")
            
            # Format the Excel file for better readability
            self._format_excel_file(filename)
            
            print(f"✅ Content appended to: {filename}")
            print(f"📊 Total entries in file: {len(new_df)}")
            
            return filename
            
        except Exception as e:
            print(f"❌ Error updating Excel file: {e}")
            return self._create_backup_file(content)
    
    def _format_excel_file(self, filename: str):
        """Format Excel file with new columns including episode title"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = load_workbook(filename)
            
            if 'Content_Log' in wb.sheetnames:
                ws = wb['Content_Log']
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Updated column widths for new columns
                column_widths = {
                    'A': 12,  # Date
                    'B': 10,  # Time
                    'C': 12,  # Day
                    'D': 20,  # Content_Type
                    'E': 80,  # Script
                    'F': 50,  # Social_Post
                    'G': 60,  # Motion_Script
                    'H': 40,  # Video_Caption
                    'I': 50,  # Episode_Title
                    'J': 15,  # Script_Length
                    'K': 12,  # Word_Count
                    'L': 12,  # News_Count
                    'M': 25,  # Market_Data
                    'N': 15   # Quality_Score
                }
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # Text wrapping for script, social post, motion script, caption, and title columns
                for row in ws.iter_rows(min_row=2):
                    if len(row) >= 9:  # Ensure we have enough columns
                        row[4].alignment = Alignment(wrap_text=True)  # Script
                        row[5].alignment = Alignment(wrap_text=True)  # Social_Post
                        row[6].alignment = Alignment(wrap_text=True)  # Motion_Script
                        row[7].alignment = Alignment(wrap_text=True)  # Video_Caption
                        row[8].alignment = Alignment(wrap_text=True)  # Episode_Title
            
            wb.save(filename)
            
        except ImportError:
            print("⚠️ openpyxl not available for advanced formatting")
        except Exception as e:
            print(f"⚠️ Could not format Excel file: {e}")
    
    def _create_backup_file(self, content: Dict[str, Any]) -> str:
        """Create timestamped backup file if main update fails"""
        date_str = datetime.now().strftime('%Y%m%d_%H%M')
        day = content['day'].lower()
        filename = f"market_content_backup_{day}_{date_str}.xlsx"
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Single row for this content
                content_data = pd.DataFrame({
                    'Date': [datetime.now().strftime('%Y-%m-%d')],
                    'Time': [datetime.now().strftime('%H:%M:%S')],
                    'Day': [content['day']],
                    'Content_Type': [content['type']],
                    'Script': [content['script']],
                    'Social_Post': [content['social_post']],
                    'Motion_Script': [content.get('motion_script', '')],
                    'Video_Caption': [content.get('video_caption', '')],
                    'Episode_Title': [content.get('episode_title', '')],
                    'News_Count': [content.get('news_count', 0)]
                })
                content_data.to_excel(writer, sheet_name='Content', index=False)
            
            return filename
        except Exception as e:
            print(f"❌ Error creating backup file: {e}")
            return ""


def main():
    print("=== Enhanced Alpha Vantage News Generator (Professional Dynamic Style) ===")
    print("📋 Checking dependencies...")
    
    # Check optional dependencies
    try:
        import yfinance
        print("✅ yfinance available")
    except ImportError:
        print("⚠️ yfinance not installed - market data will be limited")
        print("   Install with: pip install yfinance")
    
    try:
        # Enable debug mode to see news filtering
        generator = ProfessionalNewsGenerator(debug_mode=True)
        print("✅ APIs initialized")
        
        content = generator.generate_content()
        
        print("\n" + "="*60)
        print("📺 GENERATED SCRIPT:")
        print("="*60)
        print(content['script'])
        
        print("\n" + "="*60)
        print("📱 SOCIAL MEDIA POST:")
        print("="*60)
        print(content['social_post'])
        
        print("\n" + "="*60)
        print("🎬 MOTION SCRIPT:")
        print("="*60)
        print(content['motion_script'])
        
        print("\n" + "="*60)
        print("📺 VIDEO CAPTION:")
        print("="*60)
        print(content['video_caption'])
        
        print("\n" + "="*60)
        print("🎯 EPISODE TITLE:")
        print("="*60)
        print(content['episode_title'])
        
        print("\n" + "="*60)
        print("📊 GENERATION SUMMARY:")
        print("="*60)
        print(f"Day: {content['day']}")
        print(f"Type: {content['type']}")
        print(f"News articles processed: {content['news_count']}")
        print(f"Script length: {len(content['script'])} characters")
        print(f"Word count: {len(content['script'].split())} words")
        print(f"Motion script length: {len(content['motion_script'])} characters")
        print(f"API calls made: {generator.call_count}")
        
        # Save to Excel
        filename = generator.save_to_excel(content)
        if filename:
            print(f"✅ Content saved to: {filename}")
        
    except Exception as e:
        print(f"❌ Fatal error: {e}")
        print("\n🔧 Troubleshooting tips:")
        print("1. Verify .env file contains ALPHA_VANTAGE_API_KEY and OPENAI_API_KEY")
        print("2. Check API key validity and quotas")
        print("3. Ensure internet connection is stable")
        print("4. Install required packages: pip install openai python-dotenv pandas openpyxl")


if __name__ == "__main__":
    main()
